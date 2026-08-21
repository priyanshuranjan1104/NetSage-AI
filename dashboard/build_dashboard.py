"""
Builds dashboard.xlsx:
  - Raw Data: cases + AI diagnosis + human verdict, joined
  - Dashboard: pivot-style formula summary by category, severity, and verdict
All summary numbers are formulas (COUNTIFS/SUMPRODUCT), not hardcoded,
so the dashboard recalculates if cases are added later.
"""
import csv
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference

BASE = Path(__file__).resolve().parent.parent

with open(f"{BASE}/cases.csv", newline="", encoding="utf-8") as f:
    cases = list(csv.DictReader(f))

with open(f"{BASE}/ai_diagnosis/ai_diagnoses.json", encoding="utf-8") as f:
    diagnoses = json.load(f)

with open(f"{BASE}/review/human_review_log.csv", newline="", encoding="utf-8") as f:
    review = {row["case_id"]: row for row in csv.DictReader(f)}

wb = Workbook()

# ---------------- Raw Data sheet ----------------
raw = wb.active
raw.title = "Raw Data"
headers = ["case_id", "category", "severity", "osi_layer", "ai_confidence", "verdict"]
raw.append(headers)
for cell in raw[1]:
    cell.font = Font(bold=True, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")

for c in cases:
    r = review[c["case_id"]]
    raw.append([c["case_id"], c["category"], c["severity"], c["osi_layer"],
                r["ai_confidence"], r["verdict"]])

for col, width in zip("ABCDEF", [10, 12, 10, 12, 14, 12]):
    raw.column_dimensions[col].width = width

n = len(cases)
last_row = n + 1  # header + n rows

# ---------------- Dashboard sheet ----------------
dash = wb.create_sheet("Dashboard")
dash["A1"] = "NetSage AI — Case Dashboard"
dash["A1"].font = Font(bold=True, size=14, name="Arial")

dash["A3"] = "Summary"
dash["A3"].font = Font(bold=True, size=12, name="Arial")
dash["A4"] = "Total cases"
dash["B4"] = f"=COUNTA('Raw Data'!A2:A{last_row})"
dash["A5"] = "AI Accepted (agreement)"
dash["B5"] = f"=COUNTIF('Raw Data'!F2:F{last_row},\"Accepted\")"
dash["A6"] = "AI Rejected (corrected)"
dash["B6"] = f"=COUNTIF('Raw Data'!F2:F{last_row},\"Rejected\")"
dash["A7"] = "AI Edited"
dash["B7"] = f"=COUNTIF('Raw Data'!F2:F{last_row},\"Edited\")"
dash["A8"] = "AI vs human agreement rate"
dash["B8"] = "=B5/B4"
dash["B8"].number_format = "0.0%"

for row in range(4, 9):
    dash[f"A{row}"].font = Font(name="Arial")
    dash[f"B{row}"].font = Font(name="Arial", bold=(row == 8))

# Issue type (category) breakdown
dash["D3"] = "Cases by issue type"
dash["D3"].font = Font(bold=True, size=12, name="Arial")
categories = sorted(set(c["category"] for c in cases))
dash["D4"] = "Category"
dash["E4"] = "Count"
for cell in (dash["D4"], dash["E4"]):
    cell.font = Font(bold=True, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
for i, cat in enumerate(categories, start=5):
    dash[f"D{i}"] = cat
    dash[f"E{i}"] = f"=COUNTIF('Raw Data'!B2:B{last_row},D{i})"
cat_last_row = 4 + len(categories)

# Severity breakdown
sev_start_row = cat_last_row + 2
dash[f"D{sev_start_row}"] = "Cases by severity"
dash[f"D{sev_start_row}"].font = Font(bold=True, size=12, name="Arial")
dash[f"D{sev_start_row+1}"] = "Severity"
dash[f"E{sev_start_row+1}"] = "Count"
for cell in (dash[f"D{sev_start_row+1}"], dash[f"E{sev_start_row+1}"]):
    cell.font = Font(bold=True, name="Arial")
    cell.fill = PatternFill("solid", fgColor="D9E1F2")
severities = ["High", "Medium", "Low"]
for i, sev in enumerate(severities, start=sev_start_row + 2):
    dash[f"D{i}"] = sev
    dash[f"E{i}"] = f"=COUNTIF('Raw Data'!C2:C{last_row},D{i})"
sev_last_row = sev_start_row + 1 + len(severities)

for col, width in zip("ABCDE", [26, 10, 4, 16, 8]):
    dash.column_dimensions[col].width = width

# ---------------- Charts ----------------
bar = BarChart()
bar.title = "Cases by Issue Type"
bar.y_axis.title = "Count"
data = Reference(dash, min_col=5, min_row=4, max_row=cat_last_row)
cats = Reference(dash, min_col=4, min_row=5, max_row=cat_last_row)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 8
dash.add_chart(bar, "G4")

pie = PieChart()
pie.title = "AI vs Human Verdict"
verdict_labels = ["Accepted", "Rejected"]
dash["D20"] = "Verdict"
dash["E20"] = "Count"
dash["D21"] = "Accepted"
dash["E21"] = "=B5"
dash["D22"] = "Rejected"
dash["E22"] = "=B6"
pdata = Reference(dash, min_col=5, min_row=20, max_row=22)
pcats = Reference(dash, min_col=4, min_row=21, max_row=22)
pie.add_data(pdata, titles_from_data=True)
pie.set_categories(pcats)
pie.width, pie.height = 12, 8
dash.add_chart(pie, "G20")

wb.save(f"{BASE}/dashboard/dashboard.xlsx")
print("Saved dashboard.xlsx")
